from datetime import timedelta

import xlsxwriter
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Sum, Func, F, Count, Q
from django.db.models.functions import TruncMonth
from users.models import Officer
from django.contrib.auth.mixins import UserPassesTestMixin
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.views.generic import DetailView, DeleteView, UpdateView, ListView
from openpyxl import *
from openpyxl.writer.excel import save_virtual_workbook
from transliterate import translit
from .forms import ProductForm, EarlyRepaymentForm
from .models import Product, Zalog_types, PrihodRashod, WorkDays
from users.models import name_user
from django.utils import timezone
from datetime import datetime
from django.contrib import messages
from django.shortcuts import redirect
from django.core.exceptions import ObjectDoesNotExist


def is_valid_queryparam(param):
    return param != '' and param is not None


class Month(Func):
    function = 'EXTRACT'
    template = '%(function)s(MONTH from %(expressions)s)'
    output_field = models.IntegerField()


def index(request):
    products = Product.objects.values('credit_user').order_by('credit_user').annotate(total_summ=Sum('summa_zayavki'))
    users = []
    data = []
    labels_bar = []
    data_bar = []
    iteration = 0
    for product in products:
        for key in product:
            iteration += 1
            if iteration % 2 == 0:
                data.append(product[key])
            else:
                users.append(product[key])
    labels = []
    for i in users:
        user = User.objects.get(id=i)
        labels.append(Officer.objects.get(user=user).name)

    bar = Product.objects.values(month=TruncMonth('date_posted')) \
        .annotate(total_summ=Sum('summa_zayavki')) \
        .order_by('date_posted')

    for b in bar:
        labels_bar.append(b['month'].strftime("%B %Y"))
        data_bar.append(b['total_summ'])

    status_list = []
    count_status_list = []

    product_count = Product.objects.count()
    status = Product.objects.values('status').annotate(count=Count('status')).order_by('status')

    for s in status:
        s['count'] = int(s['count'] / product_count * 100)

    prihod = PrihodRashod.objects.all().filter(
        date__month=WorkDays.objects.all().order_by('-day')[0].day.strftime("%m")).values('type').annotate(
        count=Sum('summa')).order_by('type')

    context = {'labels': labels,
               'data': data,
               'bar': bar,
               'labels_bar': labels_bar,
               'data_bar': data_bar,
               'status_list': status_list,
               'count_status_list': count_status_list,
               'status': status,
               'prihod': prihod}

    return render(request, 'journal/dashboard.html', context)


def product_list(request):
    products = Product.objects.all().order_by('-nomer_bileta')
    try:
        single_product = Product.objects.all()[0]
    except:
        single_product = {}
    product_per_page = 10
    paginator = Paginator(products, product_per_page)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'product_list.html', {'products': page_obj, 'single_product': single_product})


def save_product_form(request, form, template_name, title):
    user = request.user
    data = dict()
    title = title
    if request.method == 'POST':
        if form.is_valid():
            obj = form.save(commit=False)
            obj.credit_user = Officer.objects.get(user=user)
            summa_zayavki = int(form.cleaned_data['summa_zayavki'])
            stavka_day = float(form.cleaned_data['stavka_day'])
            srok_kredita = int(form.cleaned_data['srok_kredita'])
            fio_klienta = form.cleaned_data['fio_klienta']
            date_posted = form.cleaned_data['date_posted']
            obj.ostatok = form.cleaned_data['summa_zayavki']

            obj.stavka_period = summa_zayavki * stavka_day / 100 * srok_kredita
            obj.fact_loan_day = 0
            obj.fact_day = 0

            obj.save()

            prihod_rashod = PrihodRashod(summa=summa_zayavki,
                                         type='Расход',
                                         date=date_posted,
                                         product=obj,
                                         comment='Выдача займа {} на сумму {} сом'.format(fio_klienta, summa_zayavki))
            prihod_rashod.save()

            data['form_is_valid'] = True
            products = Product.objects.all()
            data['html_product_list'] = render_to_string('includes/partial_product_list.html', {
                'products': products
            })
        else:
            data['form_is_valid'] = False
    context = {'form': form,
               'title': title}
    data['html_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)


def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
    else:
        form = ProductForm()
    title = 'Выдача займа'
    return save_product_form(request, form, 'includes/partial_product_create.html', title)


class DelayListView(ListView):
    queryset = Product.objects.all().filter(status='Просрочен').order_by('-nomer_bileta')
    template_name = 'journal/delays.html'
    context_object_name = 'products'
    paginate_by = 10

    def get_context_data(self, *args, **kwargs):
        context = super(DelayListView, self).get_context_data(**kwargs)
        context['title'] = 'Активные заявки'
        return context


def prihod_rashod(request):
    queryset = PrihodRashod.objects.all()
    date_min_prihod = request.GET.get('date_min_prihod')
    date_max_prihod = request.GET.get('date_max_prihod')

    if is_valid_queryparam(date_min_prihod):
        queryset = queryset.filter(date__gte=date_min_prihod)

    if is_valid_queryparam(date_max_prihod):
        queryset = queryset.filter(date__lt=date_max_prihod)

    prihod_summ = queryset.values('type').order_by('type').annotate(total_summ=Sum('summa'))

    context = {'queryset': queryset,
               'date_min_prihod': date_min_prihod,
               'date_max_prihod': date_max_prihod,
               'prihod_summ': prihod_summ}
    return render(request, 'journal/prihod_rashod.html', context)


def SearchFilterView(request):
    query_set = Product.objects.all().order_by('-nomer_bileta')
    search_contains_query = request.GET.get('search_contains')
    search_exact_query = request.GET.get('search_exact')
    sum_Count_Min = request.GET.get('sum_Count_Min')
    sum_Count_Max = request.GET.get('sum_Count_Max')
    date_min = request.GET.get('date_min')
    date_max = request.GET.get('date_max')
    status = request.GET.get('status')
    date_min_end = request.GET.get('date_min_end')
    date_max_end = request.GET.get('date_max_end')
    officer = request.GET.get('officer')
    main_search = request.GET.get('main_search')
    title = 'Результаты поиска'

    if is_valid_queryparam(main_search):
        query_set = query_set.filter(
            Q(fio_klienta__icontains=main_search.title()) | Q(nomer_bileta__iexact=main_search.title()))

    if is_valid_queryparam(search_contains_query):
        query_set = query_set.filter(fio_klienta__icontains=search_contains_query.title())

    if is_valid_queryparam(search_exact_query):
        query_set = query_set.filter(nomer_bileta__iexact=search_exact_query)

    if is_valid_queryparam(sum_Count_Min):
        query_set = query_set.filter(summa_zayavki__gte=sum_Count_Min)

    if is_valid_queryparam(sum_Count_Max):
        query_set = query_set.filter(summa_zayavki__lt=sum_Count_Max)

    if is_valid_queryparam(date_min):
        query_set = query_set.filter(date_posted__gte=date_min)

    if is_valid_queryparam(date_max):
        query_set = query_set.filter(date_posted__lt=date_max)

    if is_valid_queryparam(status) and status != 'Выберите статус...':
        query_set = query_set.filter(status__iexact=status)

    if is_valid_queryparam(officer) and officer != 'Выберите сотрудника...':
        query_set = query_set.filter(credit_user__name=officer)

    if is_valid_queryparam(date_min_end):
        query_set = query_set.filter(date_plan_pay__gte=date_min_end)

    if is_valid_queryparam(date_max_end):
        query_set = query_set.filter(date_plan_pay__lt=date_max_end)

    qs = query_set.values_list('id', flat=True)
    excel_list = []
    for i in qs:
        excel_list.append(i)

    request.session['excel_list'] = excel_list

    product_per_page = 10
    paginator = Paginator(query_set, product_per_page)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'queryset': page_obj,
        'title': title,
    }

    return render(request, 'journal/search_form.html', context)


class ProductDetailView(DetailView):
    model = Product

    def test(self):
        user = self.get_object().credit_user
        username = name_user(user)  # Данный метод описан в файле users\models.py

        if self.request.user == username:
            return True
        return False

    def get_context_data(self, *args, **kwargs):
        context = super(ProductDetailView, self).get_context_data(**kwargs)
        context['title'] = 'Подробно'
        context['products'] = Zalog_types.objects.all()
        return context


def excel_report(request, pk):
    client = get_object_or_404(Client, pk=pk)
    file = 'D:/Python/changan_journal/keylist.xlsx'

    key_book = load_workbook(file)
    worksheet = key_book['list']
    worksheet_graphic = key_book['graphic']

    customer_name = worksheet.cell(8, 5)
    customer_name.value = client.fio_klienta

    officer_name = worksheet.cell(10, 5)
    officer_name.value = client.credit_user.name

    date_pay = [1, 5, 10, 15, 20, 25]
    day = int(client.date_posted.strftime('%d'))
    if day not in date_pay:
        pay_day = min(date_pay, key=lambda x: abs(x - day))
        delta = pay_day - day
        count_date = client.date_posted + timedelta(days=delta)
    else:
        count_date = client.date_posted

    date_count = worksheet_graphic.cell(7, 3)
    date_count.value = count_date

    date = worksheet_graphic.cell(4, 4)
    date.value = client.date_posted

    commission = worksheet.cell(24, 5)
    comm = client.get_commission() / 100
    commission.value = comm

    obnal = worksheet.cell(25, 5)
    ob = client.product.commission_cash / 100
    obnal.value = ob

    zalog = worksheet.cell(35, 5)
    zalog.value = client.zalog

    sum = worksheet_graphic.cell(1, 4)
    sum.value = client.summa_zayavki

    stavka = worksheet_graphic.cell(2, 4)
    stavka.value = client.get_interest_rate()

    srok = worksheet_graphic.cell(3, 4)
    srok.value = client.srok_kredita

    key_book.save(file)

    def splitpart(value):
        if len(value.split()) == 3:
            return f'{value.split()[0]}{value.split()[1][0]}.{value.split()[2][0]}.'
        elif len(value.split()) == 2:
            return f'{value.split()[0]}{value.split()[1][0]}.'
        else:
            return f'{value.split()[0]}'

    file_name = splitpart(client.fio_klienta)
    file_name = translit(file_name, "ru", reversed=True)

    response = HttpResponse(save_virtual_workbook(key_book), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment; filename={}.xlsx".format(file_name)

    return response


class ProductDeleteView(UserPassesTestMixin, DeleteView):
    model = Product
    success_url = '/products'

    def get_context_data(self, *args, **kwargs):
        context = super(ProductDeleteView, self).get_context_data(**kwargs)
        context['title'] = 'Удаление заявки'
        context['products'] = Zalog_types.objects.all()
        return context

    def test_func(self):
        user = self.get_object().credit_user
        username = name_user(user)  # Данный метод описан в файле users\models.py

        if self.request.user == username:
            return True
        return False


class ProductUpdateView(UserPassesTestMixin, UpdateView):
    model = Product
    fields = ['date_posted',
              'nomer_bileta',
              'fio_klienta',
              'summa_zayavki',
              'status',
              'stavka_day',
              'srok_kredita',
              'comment',
              'credit_user']

    def get_context_data(self, *args, **kwargs):
        context = super(ProductUpdateView, self).get_context_data(**kwargs)
        context['title'] = 'Измнение заявки'
        context['products'] = Product.objects.all()
        return context

    def test_func(self):
        user = self.get_object().credit_user
        username = name_user(user)

        if self.request.user == username:
            return True
        return False


def excel_search(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename=Filter_loans.xlsx"
    context = {
        'in_memory': True,
        'remove_timezone': True
    }
    background = '#ebfcff'
    border = 1

    def add_custom_format(format):
        format.set_bg_color(background)
        format.set_border(border)
        format.set_bold()

        return format

    book = xlsxwriter.Workbook(response, context)
    sheet = book.add_worksheet('Займы')
    sheet.set_column(0, 0, 6)
    sheet.set_column(1, 1, 12)
    sheet.set_column(2, 2, 35)
    sheet.set_column(3, 3, 13)
    sheet.set_column(4, 4, 16)
    sheet.set_column(5, 5, 12)
    sheet.set_column(6, 6, 22)
    sheet.set_column(7, 7, 16)
    sheet.set_column(8, 8, 16)
    sheet.set_column(9, 9, 19)
    sheet.set_column(10, 10, 30)
    sheet.set_column(11, 11, 20)
    sheet.set_column(12, 12, 16)
    sheet.set_column(13, 13, 20)
    sheet.set_column(14, 14, 16)
    sheet.set_column(15, 15, 20)
    sheet.set_zoom(80)

    number_format = book.add_format({'num_format': '№#####0'})
    add_custom_format(number_format)

    day_format = book.add_format({'num_format': '#####0'})
    add_custom_format(day_format)

    percent_format = book.add_format({'num_format': '#,##.#0 %'})
    add_custom_format(percent_format)

    money_format = book.add_format({'num_format': '#,###,##0 сом'})
    add_custom_format(money_format)

    date_format = book.add_format({'num_format': 'dd.mm.yyyy'})
    add_custom_format(date_format)

    style_format = book.add_format()
    add_custom_format(style_format)

    col = 0
    row = 0
    sheet.write(0, col, Product._meta.get_field("nomer_bileta").verbose_name, style_format)
    sheet.write(row, col + 1, Product._meta.get_field("date_posted").verbose_name, style_format)
    sheet.write(row, col + 2, Product._meta.get_field("fio_klienta").verbose_name, style_format)
    sheet.write(row, col + 3, Product._meta.get_field("summa_zayavki").verbose_name, style_format)
    sheet.write(row, col + 4, Product._meta.get_field("status").verbose_name, style_format)
    sheet.write(row, col + 5, Product._meta.get_field("srok_kredita").verbose_name, style_format)
    sheet.write(row, col + 6, Product._meta.get_field("stavka_day").verbose_name, style_format)
    sheet.write(row, col + 7, Product._meta.get_field("ostatok").verbose_name, style_format)
    sheet.write(row, col + 8, Product._meta.get_field("stavka_period").verbose_name, style_format)
    sheet.write(row, col + 9, Product._meta.get_field("expected_stavka_day").verbose_name, style_format)
    sheet.write(row, col + 10, Product._meta.get_field("fact_day").verbose_name, style_format)
    sheet.write(row, col + 11, Product._meta.get_field("fact_loan_day").verbose_name, style_format)
    sheet.write(row, col + 12, Product._meta.get_field("date_plan_pay").verbose_name, style_format)
    sheet.write(row, col + 13, Product._meta.get_field("date_fact_pay").verbose_name, style_format)
    sheet.write(row, col + 14, Product._meta.get_field("comment").verbose_name, style_format)
    sheet.write(row, col + 15, Product._meta.get_field("credit_user").verbose_name, style_format)
    row = 1
    col = 0
    excel_list = request.session['excel_list']
    for a in excel_list:
        queryset = Product.objects.get(id=int(a))
        sheet.write(row, col, queryset.nomer_bileta, number_format)
        sheet.write(row, col + 1, queryset.date_posted, date_format)
        sheet.write(row, col + 2, queryset.fio_klienta, style_format)
        sheet.write(row, col + 3, queryset.summa_zayavki, money_format)
        sheet.write(row, col + 4, queryset.status, style_format)
        sheet.write(row, col + 5, queryset.srok_kredita, day_format)
        sheet.write(row, col + 6, queryset.stavka_day / 100, percent_format)
        sheet.write(row, col + 7, queryset.ostatok, money_format)
        sheet.write(row, col + 8, queryset.stavka_period, money_format)
        sheet.write(row, col + 9, queryset.expected_stavka_day, money_format)
        sheet.write(row, col + 10, queryset.fact_day, money_format)
        sheet.write(row, col + 11, queryset.fact_loan_day, number_format)
        sheet.write(row, col + 12, queryset.date_plan_pay, date_format)
        sheet.write(row, col + 13, queryset.date_fact_pay, date_format)
        sheet.write(row, col + 14, queryset.comment, style_format)
        sheet.write(row, col + 15, queryset.credit_user.name, style_format)
        row = row + 1
    book.close()

    return response


def early_repayment(request, pk):
    summ = float(request.GET.get('summ'))
    try:
        product = Product.objects.get(pk=pk)
        new_summ = product.itogo_k_vyplate - summ
        product.itogo_k_vyplate = new_summ
        if new_summ == 0:  # Если полное досрочное
            product.ostatok = 0
            product.fact_day = 0
            product.itogo_k_vyplate = 0
            product.date_fact_pay = timezone.now().date()
            days = product.date_fact_pay - product.date_posted
            product.fact_loan_day = days.days
            if product.status == 'На продаже':
                product.status = 'Реализован'
            else:
                product.status = 'Выкуп'

            product.expected_stavka_day = product.ostatok * product.stavka_day / 100
            product.stavka_period = product.expected_stavka_day * (product.srok_kredita - product.fact_loan_day)

            product.save()

            prihod_rashod = PrihodRashod(summa=summ,
                                         type='Приход',
                                         date=product.date_fact_pay,
                                         product=product,
                                         comment='Полное досрочное погашение {} на сумму {}'.format(product.fio_klienta,
                                                                                                    summ))
            prihod_rashod.save()

            messages.success(request, 'Успешно проведено полное досрочное погашение')
        elif new_summ > 1:  # Сумма меньше требуемой для полного досрочного
            if product.fact_day <= summ:  # Если сумма покрывает ожидаемый доход по процентам
                product.itogo_k_vyplate = product.itogo_k_vyplate - summ
                summ = summ - product.fact_day

                product.fact_day = 0

                product.ostatok = product.ostatok - summ

                product.expected_stavka_day = product.ostatok * product.stavka_day / 100
                product.stavka_period = product.expected_stavka_day * (product.srok_kredita - product.fact_loan_day)

                product.save()

                prihod_rashod = PrihodRashod(summa=summ,
                                             type='Приход',
                                             date=timezone.now(),
                                             product=product,
                                             comment='Частично досрочное погашение {} на сумму {}'.format(
                                                 product.fio_klienta, summ))
                prihod_rashod.save()
                messages.success(request, 'Успешно проведено частично-досрочное погашение')
            else:
                product.itogo_k_vyplate = product.itogo_k_vyplate - summ
                product.fact_day = product.fact_day - summ
                product.expected_stavka_day = product.ostatok * product.stavka_day / 100
                product.stavka_period = product.expected_stavka_day * (product.srok_kredita - product.fact_loan_day)
                product.save()
                prihod_rashod = PrihodRashod(summa=summ,
                                             type='Приход',
                                             date=timezone.now(),
                                             product=product,
                                             comment='Частично досрочное погашение {} на сумму {}'.format(
                                                 product.fio_klienta, summ))
                prihod_rashod.save()
                messages.success(request, 'Успешно проведено частично-досрочное погашение по процентам')
        else:
            messages.warning(request, 'Сумма для досрочного погашение больше необходимой суммы')
    except ObjectDoesNotExist:
        messages.error(request, 'Займ не найден')
    return redirect('product_detail', pk=product.pk)


def new_day(request):
    products = Product.objects.all()
    calc_proc = 0
    today = WorkDays.objects.all().order_by('-day')[0]
    next_day = today.day + timezone.timedelta(days=1)
    for product in products:
        if product.status != 'Выкуп' or product.status != 'Реализован':
            product.fact_day = product.fact_day + product.expected_stavka_day
            calc_proc += product.expected_stavka_day
            product.fact_loan_day += 1
            product.save()
    workday = WorkDays(day=next_day, is_calculated=True, calc_sum=calc_proc)
    workday.save()
    messages.success(request, 'Начисление процентов осуществлено успешно')
    return redirect('product_list')


def prolongation(request, pk):
    days = int(request.GET.get('days'))
    try:
        product = Product.objects.get(pk=pk)
        product.srok_kredita = product.srok_kredita + days
        product.status = 'Продлен'
        product.stavka_period = product.expected_stavka_day * (product.srok_kredita - product.fact_loan_day)
        product.save()

        messages.success(request, 'Займ успешно продлен')
    except ObjectDoesNotExist:
        messages.error(request, 'Займ не найден')
    return redirect('product_detail', pk=product.pk)


def for_sale(request, pk):
    try:
        product = Product.objects.get(pk=pk)
        product.stavka_day = 0.3
        product.status = 'На продаже'

        product.save()

        messages.success(request, 'Залог в статусе выставлен на продажу')
    except ObjectDoesNotExist:
        messages.error(request, 'Займ не найден')
    return redirect('product_detail', pk=product.pk)
